import subprocess
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, QPushButton, QVBoxLayout, 
                             QHBoxLayout, QWidget, QTableView, QMessageBox, QSizePolicy, QDialog)
from PyQt5.QtCore import QAbstractTableModel, Qt
from PyQt5.QtGui import QColor
import pandas as pd
from openpyxl import load_workbook
import psycopg2
from config import config
import sys
import os
from dashboard import Dashboard

class PandasModel(QAbstractTableModel):
    def __init__(self, df=pd.DataFrame()):
        super(PandasModel, self).__init__()
        self.df = df

    def rowCount(self, parent=None):
        return len(self.df)

    def columnCount(self, parent=None):
        return len(self.df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.DisplayRole:
            return str(self.df.iloc[index.row(), index.column()])
        if role == Qt.BackgroundRole and index.column() == 2:
            return QColor(220, 220, 255)
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self.df.columns[section]
            else:
                return str(section + 1)
        return None

    def setData(self, index, value, role=Qt.EditRole):
        if index.isValid() and role == Qt.EditRole:
            col_name = self.df.columns[index.column()]  # obtenir le nom de la colonne à partir de l'index
            if col_name == 'Commentaire':
                self.df.iat[index.row(), index.column()] = value
                self.dataChanged.emit(index, index)
                return True
            elif col_name == 'Travail':
                try:
                    # analyser le temps de 'Travail' au format hh:mm
                    new_time = pd.to_timedelta(value + ':00') 
                    # convertir timedelta au format hh:mm
                    time_str = str(new_time).split()[2][:5]  # extraire 'hh:mm'
                    self.df.iat[index.row(), index.column()] = time_str
                    # mettre à jour le temps cumulé à partir de cette ligne
                    self.update_cumulative_travail(start_index=index.row())
                    self.dataChanged.emit(index, index)
                    return True
                except ValueError:
                    return False
            elif col_name == 'Date':
                try:
                    # analyser 'Date' au format yyyy-mm-dd
                    new_date = pd.to_datetime(value, format='%Y-%m-%d', errors='coerce')
                    if pd.notna(new_date):
                        self.df.iat[index.row(), index.column()] = new_date.strftime('%Y-%m-%d')
                        return True
                    else:
                        return False
                except Exception as e:
                    print(f"Error updating Date column: {e}")
                    return False
        return False

    def update_cumulative_travail(self, start_index=0):
        cumulative_time = pd.Timedelta(0)  # initialiser le temps cumulé

        for index, row in self.df.iterrows():
            if index < start_index:
                # ignorer les lignes avant l'index de départ
                cumulative_time = pd.to_timedelta(row['Travail Cumulée'] + ':00')
                continue

            if row['Travail'] == 'Abs':
                self.df.at[index, 'Travail Cumulée'] = cumulative_time
            else:
                try:
                    travail = pd.to_timedelta(row['Travail'] + ':00')
                    cumulative_time += travail
                    self.df.at[index, 'Travail Cumulée'] = cumulative_time
                except ValueError:
                    self.df.at[index, 'Travail Cumulée'] = cumulative_time

        # formater la colonne 'Travail Cumulée' en 'hh:mm'
        self.df['Travail Cumulée'] = self.df['Travail Cumulée'].apply(lambda x: str(x).split()[-1] if pd.notna(x) else '00:00')


    def flags(self, index):
        if not index.isValid():
            return Qt.ItemIsEnabled
        if self.df.columns[index.column()] in ['Commentaire', 'Travail', 'Date']:
            return Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable

    def get_dataframe(self):
        return self.df

class ExcelFileHandler:
    def __init__(self):
        self.df = pd.DataFrame()

    def connect_db(self):
        try:
            params = config()
            connection = psycopg2.connect(**params)
            print("Database connection successful")
            return connection
        except Exception as error:
            print(f"Database connection error: {error}")
            return None

    def load_excel(self, file_path):
        if os.path.exists(file_path):
            try:
                if file_path.endswith('.xlsx'):
                    df = pd.read_excel(file_path, engine='openpyxl')
                elif file_path.endswith('.xls'):
                    df = pd.read_excel(file_path, engine='xlrd')
                else:
                    print(f"Unsupported file format: {file_path}")
                    return None

                print(f"File loaded: {file_path}")
                print("Initial DataFrame:")
                print(df.head())  # afficher les premières lignes pour vérifier les données

                if all(col in df.columns for col in ['Entrée.', 'Sortie.', 'Nom.']):
                    extracted_data = []
                    cumulative_diff = pd.Timedelta(0)
                    formatted_cum_diff = '00:00:00'

                    for _, row in df.iterrows():
                        Entrée = pd.to_datetime(row['Entrée.'], errors='coerce')
                        Sortie = pd.to_datetime(row['Sortie.'], errors='coerce')
                        Nom = row['Nom.']
                        Date = pd.to_datetime(row['Date.'], errors='coerce')

                        print(f"Processing row: Entrée={Entrée}, Sortie={Sortie}, Nom={Nom}")  

                        if pd.notna(Entrée) and pd.notna(Sortie):
                            time_diff = Sortie - Entrée
                            total_seconds = int(time_diff.total_seconds())
                            diff_hours = total_seconds // 3600
                            diff_minutes = (total_seconds % 3600) // 60
                            diff_seconds = total_seconds % 60
                            formatted_diff = f"{diff_hours:02}:{diff_minutes:02}:{diff_seconds:02}"

                            cumulative_diff += time_diff
                            cumulative_total_seconds = int(cumulative_diff.total_seconds())
                            cum_hours = cumulative_total_seconds // 3600
                            cum_minutes = (cumulative_total_seconds % 3600) // 60
                            cum_seconds = cumulative_total_seconds % 60
                            formatted_cum_diff = f"{cum_hours:02}:{cum_minutes:02}:{cum_seconds:02}"

                            extracted_data.append({
                                'Entrée': Entrée.strftime('%H:%M:%S'),
                                'Sortie': Sortie.strftime('%H:%M:%S'),
                                'Nom': Nom,
                                'Date': Date.strftime('%Y-%m-%d'),
                                'Travail': formatted_diff,
                                'Travail Cumulée': formatted_cum_diff,
                                'Commentaire': ''  
                            })
                        else:
                            extracted_data.append({
                                'Entrée': 'Abs',
                                'Sortie': 'Abs',
                                'Nom': Nom,
                                'Date': 'Abs',
                                'Travail': 'Abs',
                                'Travail Cumulée': formatted_cum_diff,
                                'Commentaire': '' 
                            })

                    extracted_df = pd.DataFrame(extracted_data)
                    reordered_columns = ['Nom', 'Date', 'Entrée', 'Sortie', 'Travail', 'Travail Cumulée', 'Commentaire']
                    extracted_df = extracted_df[reordered_columns]

                    print("Extracted and converted rows:")
                    print(extracted_df)

                    self.df = extracted_df 

                    return extracted_df
                else:
                    print("Required columns are not present in the DataFrame.")
                    return None
            except Exception as e:
                print(f"Error loading file {file_path}: {e}")
                return None
        else:
            print(f"The file {file_path} does not exist.")
            return None 

    def save_excel(self, df, output_file_path):
        try:
            df.to_excel(output_file_path, index=False)
            wb = load_workbook(output_file_path)
            ws = wb.active

            column_widths = {
                1: 30,  
                2: 20,  
                3: 20,  
                4: 20,  
                5: 20,  
                6: 30,  
                7: 50   
            }

            for col, width in column_widths.items():
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

            wb.save(output_file_path)

            print(f"DataFrame saved as {output_file_path} with adjusted column widths")
        except Exception as e:
            print(f"Error saving DataFrame as {output_file_path}: {e}")

    def insert_to_db(self, df):
        connection = self.connect_db()

        if connection is None:
            print("Failed to connect to the database.")
            return

        try:
            cursor = connection.cursor()
            insert_query = """
                INSERT INTO dbbi (Nom, Date, Travail, Travail_cumule)
                VALUES (%s, %s, %s, %s)
            """
            check_query = """
                SELECT COUNT(*) FROM dbbi WHERE Nom = %s AND Date = %s
            """

            for _, row in df.iterrows():
                nom = row['Nom'] if row['Nom'] != 'Abs' else None
                date = row['Date'] if row['Date'] != 'Abs' else None
                travail = row['Travail'] if row['Travail'] != 'Abs' else None
                travail_cumulee = row['Travail Cumulée'] if row['Travail Cumulée'] != 'Abs' else None
                
              # si Nom ou Date est None, ignorer l'insertion
                if nom is None or date is None:
                    print(f"Skipping row with missing Nom or Date: {row}")
                    continue

                print(f"Checking for existing data: (Nom: {nom}, Date: {date})")
                
                try:
                    cursor.execute(check_query, (nom, date))
                    exists = cursor.fetchone()[0]

                    if exists:
                        print(f"Data already exists: (Nom: {nom}, Date: {date})")
                        continue

                    print(f"Inserting data: (Nom: {nom}, Date: {date}, Travail: {travail}, Travail_cumulee: {travail_cumulee})")
                    cursor.execute(insert_query, (nom, date, travail, travail_cumulee))
                except Exception as e:
                    print(f"Error executing query with data (Nom: {nom}, Date: {date}, Travail: {travail}, Travail_cumulee: {travail_cumulee}): {e}")
                    connection.rollback()
                    return

            connection.commit()
            print("Data inserted successfully into the database.")
        except Exception as e:
            print(f"Error inserting data: {e}")
        finally:
            cursor.close()
            connection.close()


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.init_ui()
        self.file_handler = ExcelFileHandler()

    def init_ui(self):
        self.load_button = QPushButton("Charger le fichier Excel")
        self.save_button = QPushButton("Enregistrez le fichier Excel")
        self.insert_button = QPushButton("Insérer dans la bd")
        self.dashboard_button = QPushButton("Tableau de bord")
        self.table_view = QTableView()

        self.load_button.clicked.connect(self.load_file_dialog)
        self.save_button.clicked.connect(self.save_file_dialog)
        self.insert_button.clicked.connect(self.insert_data_to_db)  
        self.dashboard_button.clicked.connect(self.open_dashboard)

        # créer une mise en page horizontale pour les boutons
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.load_button)
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.insert_button)  
        button_layout.addWidget(self.dashboard_button)
        button_layout.setAlignment(Qt.AlignCenter)

        # définir les tailles des boutons
        self.load_button.setFixedSize(300, 70)
        self.save_button.setFixedSize(300, 70)
        self.insert_button.setFixedSize(300, 70) 
        self.dashboard_button.setFixedSize(300, 70)


        self.table_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout = QVBoxLayout()
        layout.addWidget(self.table_view)
        layout.addLayout(button_layout)  

        layout.setStretch(0, 1)  
        layout.setStretch(1, 0)  

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                border: none;
                color: white;
                padding: 15px 30px;
                text-align: center;
                text-decoration: none;
                display: inline-block;
                font-size: 18px;
                margin: 10px;
                transition-duration: 0.4s;
                cursor: pointer;
                border-radius: 8px;
            }
            
            QPushButton:hover {
                background-color: #45a049; /* Darker green */
            }
            
            QTableView::item {
                font-weight: bold;  /* Ensure table items are bold */
            }
            
            QHeaderView::section {
                font-weight: bold;  /* Ensure header text is bold */
                font-size: 16px;   /* Adjust header font size */
            }
            
            QTabWidget::pane {
                border: 1px solid #C2C7CB;
                border-radius: 8px;
                background-color: #F5F5F5;
                margin-top: 12px;
            }
            
            QTabBar::tab {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: -1px;
            }
            
            QTabBar::tab:selected {
                background-color: #2E8B57;
            }
        """)

    def load_file_dialog(self):
        options = QFileDialog.Options()
        file, _ = QFileDialog.getOpenFileName(self, "Charger le fichier Excel", "", "Fichiers Excel (*.xls *.xlsx)", options=options)
        if file:
            extracted_df = self.file_handler.load_excel(file)
            if extracted_df is not None:
                self.update_table_view(extracted_df)
            else:
                self.show_error_message("Échec du chargement du fichier Excel. Vérifiez le format du fichier et les colonnes requises.")

    def save_file_dialog(self):
        options = QFileDialog.Options()
        file, _ = QFileDialog.getSaveFileName(self, "Enregistrez le fichier Excel", "", "Fichiers Excel (*.xlsx)", options=options)
        if file:
            model = self.table_view.model()
            if model is not None:
                df = model.get_dataframe()
                self.file_handler.save_excel(df, file)

    def insert_data_to_db(self):
        model = self.table_view.model()
        if model is not None:
            df = model.get_dataframe()
            self.file_handler.insert_to_db(df)

    def open_dashboard(self):
        # cacher la fenêtre principale
        self.hide()
        dashboard = Dashboard()
        self.setCentralWidget(dashboard)
        # créer une instance de CourriersWidget
        self.showMaximized() 
        
    def update_table_view(self, df):
        model = PandasModel(df)
        self.table_view.setModel(model)

        self.table_view.setColumnWidth(0, 150)  # Nom
        self.table_view.setColumnWidth(1, 150)  # Date
        self.table_view.setColumnWidth(2, 150)  # Entrée
        self.table_view.setColumnWidth(3, 150)  # Sortie
        self.table_view.setColumnWidth(4, 200)  # Travail 
        self.table_view.setColumnWidth(5, 200)  # Travail Cumulée
        self.table_view.setColumnWidth(6, 845)  # Commentaire

    def show_error_message(self, message):
        QMessageBox.critical(self, "Erreur", message)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.showMaximized()
    sys.exit(app.exec_())
